"""
FR24 LOG VERIFICATION (replay-aware)

Matches manual flight-log entries (xlsx workbooks) against OCR screenshot
observations, attributing each observation to its TRUE flight date:

  - If the OCR raw excerpt contains an FR24 replay-bar date ("Sun, Sep 28, 2025"),
    that date is the flight date — the screenshot filename date is only the
    capture date. Replays of flights older than the log period are flagged,
    not dropped (no date floor: a Jan 2025 replay viewed in Oct 2025 is real).
  - A replay-bar date later than the capture date is OCR noise and is ignored.
  - Live views (no replay bar) keep the filename date.

Replaces the filename-date matching that produced systematic one-to-multi-day
date offsets (see spiderweb-pr/outputs/coverage_audit/P2_REVIEW.md, 2026-06-03
audit: 47 of 79 "unconfirmed" log entries were replay-date artifacts). This
module was written in spiderweb-pr before the fr24 package moved here, hence
the cross-repo reference.

Outputs:
  - log_verify_confirmed.csv      log entries with >=1 same-tail obs on true date
  - log_verify_unconfirmed.csv    log entries with no matching obs
  - log_verify_pre_log_replays.csv  replay obs whose flight date precedes the log period
  - log_verify_summary.json

Usage:
    python3 -m fr24.log_verify \
        --logs "data/Flight Log 2025.xlsx" "data/manual_logs/Flight Log 2026 (seed).xlsx" \
        --observations outputs/aircraft_observations.csv \
        --output-dir outputs/log_verify
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

VERIFIER_VERSION = "fr24_log_verify_v0.1.0"

MONTHS = {
    "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12,
}

# FR24 replay bar, e.g. "Sun, Sep 28, 2025" / "WED, SEP 3 2025" (OCR-tolerant).
REPLAY_BAR_RE = re.compile(
    r"\b(?:MON|TUE|WED|THU|FRI|SAT|SUN)[A-Z]*\.?,?\s+"
    r"(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)[A-Z]*\.?\s+"
    r"(\d{1,2}),?\s+(\d{4})\b"
)

TAIL_NOISE_RE = re.compile(r"\(.*?\)")
TAIL_KEEP_RE = re.compile(r"[^A-Z0-9-]")


def normalize_tail(raw) -> Optional[str]:
    """Normalize "N196DM (“BlueBoy”)" → "N196DM". Returns None when empty."""
    if raw is None:
        return None
    cleaned = TAIL_KEEP_RE.sub("", TAIL_NOISE_RE.sub("", str(raw).upper()))
    return cleaned or None


def extract_replay_date(raw_excerpt: str, capture_date: date) -> Optional[date]:
    """Return the replay-bar flight date, or None for live views / noise.

    No historical floor is applied: replays may reference flights arbitrarily
    far back. Dates later than the capture date are rejected as OCR noise.
    """
    if not raw_excerpt:
        return None
    match = REPLAY_BAR_RE.search(raw_excerpt.upper())
    if not match:
        return None
    month, day, year = MONTHS[match.group(1)], int(match.group(2)), int(match.group(3))
    try:
        candidate = date(year, month, day)
    except ValueError:
        return None
    if candidate > capture_date:
        return None
    return candidate


def true_flight_date(filename_ts: str, raw_excerpt: str) -> Tuple[date, bool]:
    """Return (flight_date, is_replay) for one observation row."""
    capture = datetime.fromisoformat(filename_ts).date()
    replay = extract_replay_date(raw_excerpt, capture)
    if replay is not None and replay != capture:
        return replay, True
    return capture, replay is not None


def load_observations(path: Path) -> Dict[Tuple[str, str], List[dict]]:
    """Index observations by (true_flight_date_iso, registration)."""
    csv.field_size_limit(10 ** 7)
    index: Dict[Tuple[str, str], List[dict]] = defaultdict(list)
    with path.open(newline="", encoding="utf-8") as handle:
        for row in csv.DictReader(handle):
            reg = normalize_tail(row.get("registration"))
            ts = row.get("filename_ts") or ""
            if not reg or not ts:
                continue
            flight_date, is_replay = true_flight_date(ts, row.get("raw_excerpt") or "")
            index[(flight_date.isoformat(), reg)].append(
                {
                    "filename": row.get("filename", ""),
                    "capture_date": ts[:10],
                    "is_replay": is_replay,
                    "identity_status": row.get("identity_status", ""),
                }
            )
    return index


def load_log_entries(paths: Iterable[Path]) -> List[dict]:
    """Flatten workbook rows to dicts. Layout: FN, UF, Date, Time, Tail, ..."""
    import openpyxl  # local import: heavy dependency

    entries: List[dict] = []
    for path in paths:
        workbook = openpyxl.load_workbook(path, read_only=True)
        for sheet_name in workbook.sheetnames:
            rows = workbook[sheet_name].iter_rows(values_only=True)
            next(rows, None)  # header
            for row in rows:
                if not row or row[2] is None:
                    continue
                day = str(row[2])[:10]
                if not re.match(r"\d{4}-\d{2}-\d{2}", day):
                    continue
                entries.append(
                    {
                        "fn": str(row[0] or ""),
                        "workbook": path.name,
                        "sheet": sheet_name,
                        "flight_date": day,
                        "tail_raw": str(row[4] or ""),
                        "tail": normalize_tail(row[4]),
                    }
                )
    return entries


def verify(
    log_paths: List[Path],
    observations_path: Path,
    log_period_start: Optional[str] = None,
) -> dict:
    """Match log entries to replay-aware observations."""
    obs_index = load_observations(observations_path)
    entries = load_log_entries(log_paths)

    if log_period_start is None and entries:
        log_period_start = min(e["flight_date"] for e in entries)

    confirmed, unconfirmed = [], []
    for entry in entries:
        key = (entry["flight_date"], entry["tail"]) if entry["tail"] else None
        hits = obs_index.get(key, []) if key else []
        record = {
            **entry,
            "obs_count": len(hits),
            "replay_obs_count": sum(1 for h in hits if h["is_replay"]),
            "evidence": hits[0]["filename"] if hits else "",
        }
        (confirmed if hits else unconfirmed).append(record)

    pre_log = [
        {"flight_date": day, "tail": reg, "obs_count": len(rows),
         "example": rows[0]["filename"]}
        for (day, reg), rows in sorted(obs_index.items())
        if log_period_start and day < log_period_start
        and any(r["is_replay"] for r in rows)
    ]

    return {
        "version": VERIFIER_VERSION,
        "log_period_start": log_period_start,
        "entries_total": len(entries),
        "confirmed": confirmed,
        "unconfirmed": unconfirmed,
        "pre_log_replays": pre_log,
    }


def _write_csv(path: Path, rows: List[dict]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Replay-aware flight-log verification")
    parser.add_argument("--logs", nargs="+", required=True, help="Flight log .xlsx paths")
    parser.add_argument("--observations", required=True, help="aircraft_observations.csv")
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--log-period-start", default=None,
                        help="ISO date; replays before this are flagged pre-log (default: earliest log entry)")
    args = parser.parse_args(argv)

    result = verify(
        [Path(p) for p in args.logs],
        Path(args.observations),
        args.log_period_start,
    )

    out = Path(args.output_dir)
    out.mkdir(parents=True, exist_ok=True)
    _write_csv(out / "log_verify_confirmed.csv", result["confirmed"])
    _write_csv(out / "log_verify_unconfirmed.csv", result["unconfirmed"])
    _write_csv(out / "log_verify_pre_log_replays.csv", result["pre_log_replays"])
    summary = {
        "version": result["version"],
        "log_period_start": result["log_period_start"],
        "entries_total": result["entries_total"],
        "confirmed": len(result["confirmed"]),
        "unconfirmed": len(result["unconfirmed"]),
        "pre_log_replay_groups": len(result["pre_log_replays"]),
    }
    (out / "log_verify_summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
    print(json.dumps(summary, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
